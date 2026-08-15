#!/usr/bin/env python3
"""Erzeugt business-plan/financial-model.xlsx — 5-Jahres-Businessplan KRONOS Eos.

Drei Szenarien (Konservativ / Base / Upside) mit echten Excel-Formeln,
Szenariovergleich, Sensitivitätsanalyse und Return-Rechnung.
Alle Annahmen sind in research/assumptions.md dokumentiert (Evidenzstufen A-D).

Aufruf:  python3 scripts/build_financial_model.py
"""

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- Parameter

YEARS = [2026, 2027, 2028, 2029, 2030, 2031]  # 2026 = Vorlauf-/Entwicklungsjahr
NY = len(YEARS)

SCEN = {
    "Konservativ": dict(
        units=[0, 8, 20, 35, 50, 65],
        list_asp=[0, 55, 57, 58, 59, 60],          # k€, blended Liste
        cogs_pct=[0, 0.60, 0.58, 0.56, 0.54, 0.52],  # % vom Listenpreis
        cons_per_machine=3.0,   # k€/Maschine/Jahr Verbrauchsmaterial
        cons_gm=0.55,
        svc_attach=0.40, svc_pct=0.10, svc_gm=0.50,
        sw_price=1.2, sw_attach=0.50, sw_gm=0.85,
        channel_share=0.25, channel_disc=0.30,      # ab Jahr 2 (2028)
        cac=12.0,               # k€ je Direkteinheit
        services=[0, 200, 300, 400, 450, 500],      # Applikationsprojekte/NRE, k€
        services_gm=0.45,
        fte=[7, 10, 12, 14, 15, 16],
        fte_cost=95.0,          # k€ Vollkosten
        dev_nonpers=[500, 550, 350, 250, 200, 200],   # k€
        cert=[0, 150, 0, 0, 0, 0],
        marketing=[80, 300, 400, 450, 450, 450],
        gna=[120, 250, 300, 330, 350, 350],
        grants=[250, 0, 0, 0, 0, 0],
        capex=[200, 500, 350, 300, 300, 300],       # inkl. Tooling, Demogeräte
        wc_pct=0.15,            # WC in % des Umsatzzuwachses (nach 30 % Anzahlungen)
        exit_multiple=1.0, investor_stake=0.60,
        ext_growth=0.15, ext_ebitda=[0.00, 0.02],   # Jahr 6/7 Terminalperspektive
    ),
    "Base": dict(
        units=[0, 12, 35, 75, 130, 200],
        list_asp=[0, 52, 60, 62, 63, 63.5],
        cogs_pct=[0, 0.58, 0.55, 0.52, 0.50, 0.48],
        cons_per_machine=5.0,
        cons_gm=0.60,
        svc_attach=0.50, svc_pct=0.11, svc_gm=0.50,
        sw_price=1.8, sw_attach=0.60, sw_gm=0.85,
        channel_share=0.25, channel_disc=0.30,
        cac=9.0,
        services=[0, 300, 500, 800, 1000, 1200],
        services_gm=0.45,
        fte=[8, 13, 19, 26, 32, 38],
        fte_cost=95.0,
        dev_nonpers=[500, 600, 400, 300, 300, 300],
        cert=[0, 150, 0, 0, 0, 0],
        marketing=[100, 350, 500, 700, 900, 1000],
        gna=[150, 300, 400, 500, 600, 700],
        grants=[250, 300, 250, 0, 0, 0],
        capex=[200, 550, 450, 450, 500, 550],
        wc_pct=0.12,
        exit_multiple=2.0, investor_stake=0.55,
        ext_growth=0.30, ext_ebitda=[0.10, 0.15],
    ),
    "Upside": dict(
        units=[0, 18, 55, 120, 220, 330],
        list_asp=[0, 52, 61, 63, 64, 65],
        cogs_pct=[0, 0.56, 0.53, 0.50, 0.47, 0.45],
        cons_per_machine=8.0,
        cons_gm=0.65,
        svc_attach=0.60, svc_pct=0.12, svc_gm=0.50,
        sw_price=2.4, sw_attach=0.70, sw_gm=0.85,
        channel_share=0.25, channel_disc=0.30,
        cac=7.0,
        services=[0, 400, 700, 1100, 1400, 1700],
        services_gm=0.45,
        fte=[9, 15, 22, 30, 40, 50],
        fte_cost=95.0,
        dev_nonpers=[500, 650, 450, 350, 350, 350],
        cert=[0, 150, 0, 0, 0, 0],
        marketing=[100, 400, 600, 900, 1100, 1300],
        gna=[150, 350, 450, 600, 700, 800],
        grants=[250, 1500, 1550, 0, 0, 0],   # inkl. EIC Accelerator 2,5 M€
        capex=[200, 600, 500, 550, 650, 750],
        wc_pct=0.10,
        exit_multiple=3.0, investor_stake=0.50,
        ext_growth=0.35, ext_ebitda=[0.18, 0.22],
    ),
}

# ---------------------------------------------------------------- Styling

F_TITLE = Font(name="Calibri", size=16, bold=True, color="0B1F3A")
F_H2 = Font(name="Calibri", size=12, bold=True, color="0B1F3A")
F_HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_BOLD = Font(name="Calibri", size=10, bold=True)
F_BASE = Font(name="Calibri", size=10)
F_NOTE = Font(name="Calibri", size=9, italic=True, color="666666")
FILL_HDR = PatternFill("solid", fgColor="0B1F3A")
FILL_SEC = PatternFill("solid", fgColor="DCE6F1")
FILL_IN = PatternFill("solid", fgColor="FFF7E6")     # Eingabezellen
FILL_KPI = PatternFill("solid", fgColor="E8F4EC")
THIN = Side(style="thin", color="B8C2D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_K = '#,##0" k€"'
NUM_K1 = '#,##0.0" k€"'
NUM_PCT = "0%"
NUM_PCT1 = "0.0%"
NUM_X = '0.0"x"'
NUM_U = "#,##0"


def sec(ws, row, text, ncols):
    ws.cell(row=row, column=1, value=text).font = F_H2
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = FILL_SEC
    return row + 1


def header_row(ws, row, labels, start_col=1):
    for i, lab in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + i, value=lab)
        cell.font = F_HDR
        cell.fill = FILL_HDR
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    return row + 1


def put_row(ws, row, label, values, fmt=NUM_K, bold=False, fill=None, start_col=2, note=None):
    c = ws.cell(row=row, column=1, value=label)
    c.font = F_BOLD if bold else F_BASE
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=v)
        cell.number_format = fmt
        cell.font = F_BOLD if bold else F_BASE
        cell.border = BORDER
        if fill:
            cell.fill = fill
    if note:
        nc = ws.cell(row=row, column=start_col + len(values), value=note)
        nc.font = F_NOTE
    return row + 1


# ---------------------------------------------------------------- Read Me

wb = Workbook()
ws = wb.active
ws.title = "Read Me"
ws.column_dimensions["A"].width = 118
rows = [
    ("KRONOS Eos — 5-Jahres-Finanzmodell (2026–2031)", F_TITLE),
    ("", None),
    ("Zweck: Businessplan für eine Low-Cost-AME-Plattform (25–100 k€) auf Basis der KRONOS-Technologie.", F_BASE),
    ("Struktur:", F_BOLD),
    ("  • Annahmen — zentrale Parameter mit Evidenzstufen (A–D) und Quellenverweis", F_BASE),
    ("  • Konservativ / Base / Upside — vollständige P&L- und Cash-Modelle (editierbare Formeln)", F_BASE),
    ("  • Szenariovergleich — Kernkennzahlen und Diagramme", F_BASE),
    ("  • Sensitivität — Wirkung der wichtigsten Treiber auf EBITDA Jahr 5 und Kapitalbedarf (Base)", F_BASE),
    ("  • Returns — Exit-Szenarien und Investor-Rendite (inkl. Terminalperspektive Jahr 7)", F_BASE),
    ("", None),
    ("Konventionen: Alle Werte in k€ (real, ohne Indexierung). 2026 ist Vorlauf-/Entwicklungsjahr.", F_BASE),
    ("Eingabezellen sind gelb hinterlegt — Änderungen fließen per Formel durch das gesamte Modell.", F_BASE),
    ("", None),
    ("Evidenzbasis: research/assumptions.md · research/sources.md · research/appendix/ (301 Findings)", F_BASE),
    ("Regenerierung: python3 scripts/build_financial_model.py", F_BASE),
    ("Leitprinzip: Substanz vor Optimismus — der Konservativ-Fall ist bewusst ein No-Go-Szenario.", F_BOLD),
    ("Stand: 15.08.2026", F_NOTE),
]
r = 1
for text, font in rows:
    if text:
        cell = ws.cell(row=r, column=1, value=text)
        if font:
            cell.font = font
    r += 1

# ---------------------------------------------------------------- Annahmen

ws = wb.create_sheet("Annahmen")
widths = [46, 16, 16, 16, 10, 44]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.cell(row=1, column=1, value="Zentrale Annahmen (Skalare) — Details je Jahr in den Szenario-Sheets").font = F_TITLE
r = 3
r = header_row(ws, r, ["Parameter", "Konservativ", "Base", "Upside", "Stufe", "Quelle / Herleitung"])

scalar_rows = [
    ("Verbrauchsmaterial je Maschine/Jahr (k€)", "cons_per_machine", None, "C/D", "Tinten-/Pastenpreise (NovaCentrix u.a.); akademische Nutzung"),
    ("Bruttomarge Verbrauchsmaterial", "cons_gm", NUM_PCT, "C", "Comps Materialgeschäft 55–75 %"),
    ("Service-Attach-Rate (Bestand)", "svc_attach", NUM_PCT, "C", "Industriestandard Laborgeräte"),
    ("Servicepreis (% vom Listenpreis p.a.)", "svc_pct", NUM_PCT, "C", "Benchmark 10–15 % p.a."),
    ("Bruttomarge Service", "svc_gm", NUM_PCT, "C", "Comps"),
    ("Software-Abo je Maschine (k€/Jahr)", "sw_price", None, "C/D", "Benchmarks Eiger/Digital Factory"),
    ("Software-Attach-Rate", "sw_attach", NUM_PCT, "C/D", ""),
    ("Bruttomarge Software", "sw_gm", NUM_PCT, "C", ""),
    ("Kanalanteil ab 2028", "channel_share", NUM_PCT, "C", "APES/NTV-Kanal existiert; Beachhead primär direkt"),
    ("Distributorenrabatt", "channel_disc", NUM_PCT, "C", "Branche 20–35 %"),
    ("CAC je Direkteinheit (k€)", "cac", None, "C", "10–20 % vom ASP, messe-zentriert"),
    ("FTE-Vollkosten (k€/Jahr)", "fte_cost", None, "B", "Standort Nürnberg, Mischkalkulation"),
    ("Working Capital (% vom Umsatzzuwachs)", "wc_pct", NUM_PCT, "C", "Lager + Forderungen − Verbindlichkeiten"),
    ("Exit-Multiple (x Umsatz Jahr 5)", "exit_multiple", NUM_X, "C", "Niche-HW-Comps: Markforged ~1,2x; Razor-Blade 2–4x"),
    ("Investorenanteil nach allen Tranchen", "investor_stake", NUM_PCT, "D", "Verhandlungsannahme, tranchiert"),
    ("Bruttomarge Applikationsservices", "services_gm", NUM_PCT, "C", "PoC-Studien/NRE — Teil des heutigen KRONOS-Modells"),
]
for label, key, fmt, stufe, src in scalar_rows:
    ws.cell(row=r, column=1, value=label).font = F_BASE
    for ci, sname in enumerate(["Konservativ", "Base", "Upside"], start=2):
        cell = ws.cell(row=r, column=ci, value=SCEN[sname][key])
        cell.fill = FILL_IN
        cell.border = BORDER
        cell.font = F_BASE
        if fmt:
            cell.number_format = fmt
    ws.cell(row=r, column=5, value=stufe).font = F_BASE
    ws.cell(row=r, column=6, value=src).font = F_NOTE
    r += 1

r += 1
ws.cell(row=r, column=1, value="Hinweis: Jahresreihen (Stückzahlen, ASP, COGS-%, FTE, Opex, Grants, Capex) stehen als gelbe Eingabezeilen in den Szenario-Sheets.").font = F_NOTE

# ---------------------------------------------------------------- Szenario-Sheets


def build_scenario_sheet(name):
    p = SCEN[name]
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 44
    for i in range(2, 2 + NY):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.column_dimensions[get_column_letter(2 + NY)].width = 52

    ws.cell(row=1, column=1, value=f"KRONOS Eos — Szenario {name}").font = F_TITLE
    r = 3
    r = header_row(ws, r, ["Position"] + [str(y) for y in YEARS] + ["Notiz"])

    def cols(row):
        return [f"{get_column_letter(2 + i)}{row}" for i in range(NY)]

    # ---- Inputs
    r = sec(ws, r, "EINGABEN (gelb = editierbar)", 2 + NY)
    r_units = r
    r = put_row(ws, r, "Verkaufte Einheiten", p["units"], NUM_U, fill=FILL_IN, note="D — Validierungsziel V1/V2")
    r_svcs_in = r
    r = put_row(ws, r, "Applikationsprojekte/NRE (Umsatz)", p["services"], NUM_K, fill=FILL_IN, note="PoC-Studien, Prozessintegration — heutiges KRONOS-Geschäft")
    r_asp = r
    r = put_row(ws, r, "Listenpreis blended (k€)", p["list_asp"], NUM_K1, fill=FILL_IN, note="Mix Eos One/Five/Max ≈ 25/55/20 %; 2027 = Pilotpreise")
    r_cogs = r
    r = put_row(ws, r, "Maschinen-COGS (% v. Liste)", p["cogs_pct"], NUM_PCT, fill=FILL_IN, note="C — BOM-These, Gate V3")
    r_fte = r
    r = put_row(ws, r, "FTE", p["fte"], NUM_U, fill=FILL_IN)
    r_dev = r
    r = put_row(ws, r, "Entwicklung (Sachkosten)", p["dev_nonpers"], NUM_K, fill=FILL_IN, note="Plattform-Derivat aus Helios-Technik")
    r_cert = r
    r = put_row(ws, r, "Zertifizierung (CE/UL)", p["cert"], NUM_K, fill=FILL_IN)
    r_mkt = r
    r = put_row(ws, r, "Marketing & Messen", p["marketing"], NUM_K, fill=FILL_IN, note="LOPEC, productronica, electronica, Formnext")
    r_gna = r
    r = put_row(ws, r, "G&A (Sachkosten)", p["gna"], NUM_K, fill=FILL_IN)
    r_grant = r
    r = put_row(ws, r, "Fördermittel (Zufluss)", p["grants"], NUM_K, fill=FILL_IN, note="ZIM/Bayern; Upside inkl. EIC 2,5 M€")
    r_capex = r
    r = put_row(ws, r, "Capex (inkl. Tooling, Demoflotte)", p["capex"], NUM_K, fill=FILL_IN)
    r += 1

    # ---- Installierte Basis
    r = sec(ws, r, "INSTALLIERTE BASIS", 2 + NY)
    r_cum = r
    vals = []
    for i in range(NY):
        col = get_column_letter(2 + i)
        prev = f"{get_column_letter(1 + i)}{r_cum}" if i else None
        vals.append(f"={prev}+{col}{r_units}" if i else f"={col}{r_units}")
    r = put_row(ws, r, "Kumulierte Basis (Jahresende)", vals, NUM_U, bold=True)
    r_avg = r
    vals = []
    for i in range(NY):
        col = get_column_letter(2 + i)
        prev_cum = f"{get_column_letter(1 + i)}{r_cum}" if i else "0"
        vals.append(f"=({prev_cum}+{col}{r_cum})/2")
    r = put_row(ws, r, "Ø aktive Basis (Jahresmittel)", vals, NUM_U)
    r_svcbase = r
    vals = ["=0"]
    for i in range(1, NY):
        vals.append(f"={get_column_letter(1 + i)}{r_cum}")
    r = put_row(ws, r, "Servicefähige Basis (Vorjahresbestand)", vals, NUM_U)
    r += 1

    # ---- Umsatz
    r = sec(ws, r, "UMSATZ", 2 + NY)
    r_mach = r
    vals = []
    for i in range(NY):
        col = get_column_letter(2 + i)
        if i <= 1:  # 2026/2027: nur direkt (Piloten)
            vals.append(f"={col}{r_units}*{col}{r_asp}")
        else:
            vals.append(
                f"={col}{r_units}*{col}{r_asp}*(1-Annahmen!{{ch}}*Annahmen!{{disc}})"
            )
    r = put_row(ws, r, "Maschinenumsatz", vals, NUM_K, note="ab 2028 Kanalmix mit Distributorenrabatt")
    r_cons = r
    vals = [f"={get_column_letter(2 + i)}{r_avg}*Annahmen!{{cons}}" for i in range(NY)]
    r = put_row(ws, r, "Verbrauchsmaterial", vals, NUM_K)
    r_svc = r
    vals = [
        f"={get_column_letter(2 + i)}{r_svcbase}*Annahmen!{{svca}}*Annahmen!{{svcp}}*{get_column_letter(2 + i)}{r_asp}"
        for i in range(NY)
    ]
    r = put_row(ws, r, "Service & Wartung", vals, NUM_K)
    r_sw = r
    vals = [f"={get_column_letter(2 + i)}{r_avg}*Annahmen!{{swa}}*Annahmen!{{swp}}" for i in range(NY)]
    r = put_row(ws, r, "Software (Aion-5X Abo)", vals, NUM_K)
    r_svcs = r
    vals = [f"={get_column_letter(2 + i)}{r_svcs_in}" for i in range(NY)]
    r = put_row(ws, r, "Applikationsprojekte/NRE", vals, NUM_K)
    r_rev = r
    vals = [
        f"=SUM({get_column_letter(2 + i)}{r_mach}:{get_column_letter(2 + i)}{r_svcs})"
        for i in range(NY)
    ]
    r = put_row(ws, r, "Gesamtumsatz", vals, NUM_K, bold=True, fill=FILL_KPI)
    r += 1

    # ---- Bruttomarge
    r = sec(ws, r, "BRUTTOERGEBNIS", 2 + NY)
    r_cogsm = r
    vals = [
        f"={get_column_letter(2 + i)}{r_units}*{get_column_letter(2 + i)}{r_asp}*{get_column_letter(2 + i)}{r_cogs}"
        for i in range(NY)
    ]
    r = put_row(ws, r, "COGS Maschinen", vals, NUM_K)
    r_gp = r
    vals = []
    for i in range(NY):
        col = get_column_letter(2 + i)
        vals.append(
            f"={col}{r_mach}-{col}{r_cogsm}"
            f"+{col}{r_cons}*Annahmen!{{consgm}}"
            f"+{col}{r_svc}*Annahmen!{{svcgm}}"
            f"+{col}{r_sw}*Annahmen!{{swgm}}"
            f"+{col}{r_svcs}*Annahmen!{{svcsgm}}"
        )
    r = put_row(ws, r, "Bruttoergebnis", vals, NUM_K, bold=True)
    r_gm = r
    vals = [
        f"=IF({get_column_letter(2 + i)}{r_rev}=0,0,{get_column_letter(2 + i)}{r_gp}/{get_column_letter(2 + i)}{r_rev})"
        for i in range(NY)
    ]
    r = put_row(ws, r, "Bruttomarge %", vals, NUM_PCT, fill=FILL_KPI)
    r += 1

    # ---- Opex
    r = sec(ws, r, "OPERATIVE KOSTEN", 2 + NY)
    r_pers = r
    vals = [f"={get_column_letter(2 + i)}{r_fte}*Annahmen!{{ftec}}" for i in range(NY)]
    r = put_row(ws, r, "Personal", vals, NUM_K)
    r_sell = r
    vals = []
    for i in range(NY):
        col = get_column_letter(2 + i)
        if i <= 1:
            vals.append(f"={col}{r_units}*Annahmen!{{cac}}")
        else:
            vals.append(f"={col}{r_units}*(1-Annahmen!{{ch}})*Annahmen!{{cac}}")
    r = put_row(ws, r, "Variabler Vertrieb (CAC direkt)", vals, NUM_K)
    r_opex = r
    vals = []
    for i in range(NY):
        col = get_column_letter(2 + i)
        vals.append(
            f"={col}{r_pers}+{col}{r_sell}+{col}{r_dev}+{col}{r_cert}+{col}{r_mkt}+{col}{r_gna}"
        )
    r = put_row(ws, r, "Opex gesamt", vals, NUM_K, bold=True)
    r += 1

    # ---- Ergebnis & Cash
    r = sec(ws, r, "ERGEBNIS & LIQUIDITÄT", 2 + NY)
    r_ebitda = r
    vals = [
        f"={get_column_letter(2 + i)}{r_gp}-{get_column_letter(2 + i)}{r_opex}+{get_column_letter(2 + i)}{r_grant}"
        for i in range(NY)
    ]
    r = put_row(ws, r, "EBITDA (inkl. Fördermittel)", vals, NUM_K, bold=True, fill=FILL_KPI)
    r_ebitdapct = r
    vals = [
        f"=IF({get_column_letter(2 + i)}{r_rev}=0,0,{get_column_letter(2 + i)}{r_ebitda}/{get_column_letter(2 + i)}{r_rev})"
        for i in range(NY)
    ]
    r = put_row(ws, r, "EBITDA-Marge %", vals, NUM_PCT)
    r_wc = r
    vals = []
    for i in range(NY):
        col = get_column_letter(2 + i)
        prev_rev = f"{get_column_letter(1 + i)}{r_rev}" if i else "0"
        vals.append(f"=MAX(0,({col}{r_rev}-{prev_rev})*Annahmen!{{wc}})")
    r = put_row(ws, r, "Δ Working Capital", vals, NUM_K)
    r_fcf = r
    vals = [
        f"={get_column_letter(2 + i)}{r_ebitda}-{get_column_letter(2 + i)}{r_capex}-{get_column_letter(2 + i)}{r_wc}"
        for i in range(NY)
    ]
    r = put_row(ws, r, "Free Cash Flow", vals, NUM_K, bold=True)
    r_cum_fcf = r
    vals = []
    for i in range(NY):
        col = get_column_letter(2 + i)
        prev = f"{get_column_letter(1 + i)}{r_cum_fcf}" if i else None
        vals.append(f"={prev}+{col}{r_fcf}" if i else f"={col}{r_fcf}")
    r = put_row(ws, r, "Kumulierter Cash Flow", vals, NUM_K, fill=FILL_KPI)
    r += 1

    # ---- KPI-Block
    r = sec(ws, r, "KENNZAHLEN", 2 + NY)
    r_need = r
    cum_range = f"{get_column_letter(2)}{r_cum_fcf}:{get_column_letter(1 + NY)}{r_cum_fcf}"
    ws.cell(row=r, column=1, value="Kapitalbedarf brutto (max. Cash-Tal, +15 % Puffer)").font = F_BOLD
    cell = ws.cell(row=r, column=2, value=f"=-MIN({cum_range})*1.15")
    cell.number_format = NUM_K
    cell.font = F_BOLD
    cell.fill = FILL_KPI
    cell.border = BORDER
    r += 1
    r_be = r
    ws.cell(row=r, column=1, value="Break-even-Jahr (EBITDA ≥ 0)").font = F_BOLD
    ebitda_range = f"{get_column_letter(2)}{r_ebitda}:{get_column_letter(1 + NY)}{r_ebitda}"
    years_range = '{' + ",".join(str(y) for y in YEARS) + '}'
    cell = ws.cell(
        row=r, column=2,
        value=f'=IF(MAX({ebitda_range})<0,"n. e.",MIN(IF({ebitda_range}>=0,{years_range})))',
    )
    cell.font = F_BOLD
    cell.fill = FILL_KPI
    cell.border = BORDER
    r += 2
    ws.cell(row=r, column=1, value="Zeilenreferenzen: siehe 'Szenariovergleich' und 'Returns'.").font = F_NOTE

    # Platzhalter in Formeln durch Annahmen-Zellbezüge ersetzen
    base_col = {"Konservativ": "B", "Base": "C", "Upside": "D"}[name]
    ann_row = {  # Zeilen im Annahmen-Sheet (Reihenfolge von scalar_rows, Start Zeile 4)
        "cons": 4, "consgm": 5, "svca": 6, "svcp": 7, "svcgm": 8,
        "swp": 9, "swa": 10, "swgm": 11, "ch": 12, "disc": 13,
        "cac": 14, "ftec": 15, "wc": 16, "mult": 17, "stake": 18,
        "svcsgm": 19,
    }
    repl = {k: f"${base_col}${v}" for k, v in ann_row.items()}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{" in cell.value:
                v = cell.value
                for k, ref in repl.items():
                    v = v.replace("{" + k + "}", ref)
                cell.value = v

    return {
        "ws": ws, "r_units": r_units, "r_asp": r_asp, "r_rev": r_rev, "r_gp": r_gp,
        "r_gm": r_gm, "r_ebitda": r_ebitda, "r_ebitdapct": r_ebitdapct,
        "r_cum": r_cum, "r_fcf": r_fcf, "r_cum_fcf": r_cum_fcf,
        "r_need": r_need, "r_be": r_be, "r_mach": r_mach, "r_cons": r_cons,
        "r_svc": r_svc, "r_sw": r_sw, "r_svcs": r_svcs, "r_opex": r_opex,
    }


refs = {name: build_scenario_sheet(name) for name in SCEN}

# ---------------------------------------------------------------- Szenariovergleich

ws = wb.create_sheet("Szenariovergleich")
ws.column_dimensions["A"].width = 44
for i in range(2, 5):
    ws.column_dimensions[get_column_letter(i)].width = 16
ws.cell(row=1, column=1, value="Szenariovergleich — Kernkennzahlen").font = F_TITLE
r = 3
r = header_row(ws, r, ["Kennzahl", "Konservativ", "Base", "Upside"])
last_col = get_column_letter(1 + NY)

cmp_rows = [
    ("Einheiten Jahr 5 (2031)", lambda n: f"={n}!{last_col}{refs[n]['r_units']}", NUM_U),
    ("Kumulierte installierte Basis 2031", lambda n: f"={n}!{last_col}{refs[n]['r_cum']}", NUM_U),
    ("Umsatz 2031", lambda n: f"={n}!{last_col}{refs[n]['r_rev']}", NUM_K),
    ("— davon wiederkehrend (Material+Service+SW)", lambda n: f"={n}!{last_col}{refs[n]['r_cons']}+{n}!{last_col}{refs[n]['r_svc']}+{n}!{last_col}{refs[n]['r_sw']}", NUM_K),
    ("Bruttomarge 2031", lambda n: f"={n}!{last_col}{refs[n]['r_gm']}", NUM_PCT),
    ("EBITDA 2031", lambda n: f"={n}!{last_col}{refs[n]['r_ebitda']}", NUM_K),
    ("EBITDA-Marge 2031", lambda n: f"={n}!{last_col}{refs[n]['r_ebitdapct']}", NUM_PCT),
    ("Kapitalbedarf brutto (inkl. Puffer)", lambda n: f"={n}!B{refs[n]['r_need']}", NUM_K),
    ("Break-even-Jahr", lambda n: f"={n}!B{refs[n]['r_be']}", NUM_U),
]
for label, fn, fmt in cmp_rows:
    ws.cell(row=r, column=1, value=label).font = F_BASE
    for ci, n in enumerate(SCEN, start=2):
        cell = ws.cell(row=r, column=ci, value=fn(n))
        cell.number_format = fmt
        cell.border = BORDER
    r += 1

# Datenblock für Diagramme (Werte per Formel aus den Szenario-Sheets)
chart_top = r + 2
ws.cell(row=chart_top - 1, column=1, value="Datenreihen (für Diagramme)").font = F_H2
ws.cell(row=chart_top, column=1, value="Jahr").font = F_BOLD
for i, y in enumerate(YEARS):
    ws.cell(row=chart_top, column=2 + i, value=y).font = F_BOLD
row_map = {}
for j, (label, key) in enumerate([
    ("Umsatz Konservativ", ("Konservativ", "r_rev")),
    ("Umsatz Base", ("Base", "r_rev")),
    ("Umsatz Upside", ("Upside", "r_rev")),
    ("EBITDA Konservativ", ("Konservativ", "r_ebitda")),
    ("EBITDA Base", ("Base", "r_ebitda")),
    ("EBITDA Upside", ("Upside", "r_ebitda")),
    ("Kum. Cash Base", ("Base", "r_cum_fcf")),
]):
    rr = chart_top + 1 + j
    row_map[label] = rr
    ws.cell(row=rr, column=1, value=label).font = F_BASE
    n, k = key
    for i in range(NY):
        col = get_column_letter(2 + i)
        cell = ws.cell(row=rr, column=2 + i, value=f"={n}!{col}{refs[n][k]}")
        cell.number_format = NUM_K

chart = LineChart()
chart.title = "Umsatzentwicklung nach Szenario (k€)"
chart.height, chart.width = 8, 20
data = Reference(ws, min_col=1, min_row=row_map["Umsatz Konservativ"], max_col=1 + NY, max_row=row_map["Umsatz Upside"])
cats = Reference(ws, min_col=2, min_row=chart_top, max_col=1 + NY)
chart.add_data(data, titles_from_data=True, from_rows=True)
chart.set_categories(cats)
ws.add_chart(chart, f"A{chart_top + 10}")

chart2 = LineChart()
chart2.title = "EBITDA nach Szenario (k€)"
chart2.height, chart2.width = 8, 20
data2 = Reference(ws, min_col=1, min_row=row_map["EBITDA Konservativ"], max_col=1 + NY, max_row=row_map["EBITDA Upside"])
chart2.add_data(data2, titles_from_data=True, from_rows=True)
chart2.set_categories(cats)
ws.add_chart(chart2, f"A{chart_top + 27}")

# ---------------------------------------------------------------- Sensitivität (python-berechnet)


def simulate(name, units_mult=1.0, asp_mult=1.0, cogs_delta=0.0, attach_mult=1.0, cac_mult=1.0):
    """Vereinfachte Python-Nachrechnung des Szenarios für die Sensitivitätstabelle."""
    p = SCEN[name]
    cum = 0.0
    prev_cum = 0.0
    prev_rev = 0.0
    cum_cash = 0.0
    min_cum_cash = 0.0
    ebitda_y5 = 0.0
    for i in range(NY):
        units = p["units"][i] * units_mult
        asp = p["list_asp"][i] * asp_mult
        cogs = min(0.95, p["cogs_pct"][i] + cogs_delta)
        ch = 0.0 if i <= 1 else p["channel_share"]
        mach_rev = units * asp * (1 - ch * p["channel_disc"])
        new_cum = prev_cum + units
        avg_base = (prev_cum + new_cum) / 2
        svc_base = prev_cum
        cons_rev = avg_base * p["cons_per_machine"] * attach_mult
        svc_rev = svc_base * p["svc_attach"] * p["svc_pct"] * asp * attach_mult
        sw_rev = avg_base * p["sw_attach"] * p["sw_price"] * attach_mult
        svcs_rev = p["services"][i]
        rev = mach_rev + cons_rev + svc_rev + sw_rev + svcs_rev
        gp = (mach_rev - units * asp * cogs
              + cons_rev * p["cons_gm"] + svc_rev * p["svc_gm"] + sw_rev * p["sw_gm"]
              + svcs_rev * p["services_gm"])
        sell = units * (1 - ch) * p["cac"] * cac_mult
        opex = (p["fte"][i] * p["fte_cost"] + sell + p["dev_nonpers"][i]
                + p["cert"][i] + p["marketing"][i] + p["gna"][i])
        ebitda = gp - opex + p["grants"][i]
        wc = max(0.0, (rev - prev_rev) * p["wc_pct"])
        fcf = ebitda - p["capex"][i] - wc
        cum_cash += fcf
        min_cum_cash = min(min_cum_cash, cum_cash)
        prev_cum = new_cum
        prev_rev = rev
        if i == NY - 1:
            ebitda_y5 = ebitda
    return ebitda_y5, -min_cum_cash * 1.15  # EBITDA Jahr 5, Kapitalbedarf brutto


ws = wb.create_sheet("Sensitivität")
ws.column_dimensions["A"].width = 44
for i in range(2, 8):
    ws.column_dimensions[get_column_letter(i)].width = 15
ws.cell(row=1, column=1, value="Sensitivitätsanalyse (Base Case)").font = F_TITLE
ws.cell(row=2, column=1, value="Wirkung einzelner Treiber auf EBITDA 2031 und Kapitalbedarf. Python-berechnet aus identischer Modelllogik; Formelmodell in den Szenario-Sheets.").font = F_NOTE

base_e, base_f = simulate("Base")
r = 4
r = header_row(ws, r, ["Treiber", "Variation", "EBITDA 2031 (k€)", "Δ EBITDA", "Kapitalbedarf (k€)", "Δ Bedarf"])

sens_specs = [
    ("Stückzahlen", "-30 %", dict(units_mult=0.7)),
    ("Stückzahlen", "+30 %", dict(units_mult=1.3)),
    ("Verkaufspreis (ASP)", "-15 %", dict(asp_mult=0.85)),
    ("Verkaufspreis (ASP)", "+15 %", dict(asp_mult=1.15)),
    ("Maschinen-COGS", "+5 pp", dict(cogs_delta=0.05)),
    ("Maschinen-COGS", "-5 pp", dict(cogs_delta=-0.05)),
    ("Attach-Umsätze", "-50 %", dict(attach_mult=0.5)),
    ("Attach-Umsätze", "+50 %", dict(attach_mult=1.5)),
    ("CAC", "+50 %", dict(cac_mult=1.5)),
    ("CAC", "-50 %", dict(cac_mult=0.5)),
]
tornado = []
for label, var, kw in sens_specs:
    e, f = simulate("Base", **kw)
    tornado.append((f"{label} {var}", e - base_e))
    put_row(ws, r, f"{label} ({var})",
            [None], NUM_K)  # label only; fill values manually below
    ws.cell(row=r, column=2, value=var).font = F_BASE
    for col, val, fmt in [(3, e, NUM_K), (4, e - base_e, NUM_K), (5, f, NUM_K), (6, f - base_f, NUM_K)]:
        cell = ws.cell(row=r, column=col, value=round(val))
        cell.number_format = fmt
        cell.border = BORDER
    r += 1
ws.cell(row=r, column=1, value="Basiswert (Base Case)").font = F_BOLD
for col, val in [(3, round(base_e)), (5, round(base_f))]:
    cell = ws.cell(row=r, column=col, value=val)
    cell.number_format = NUM_K
    cell.font = F_BOLD
    cell.fill = FILL_KPI
    cell.border = BORDER
r += 2

# Tornado-Daten + Chart
ws.cell(row=r, column=1, value="Tornado: Δ EBITDA 2031 (k€)").font = F_H2
r += 1
tornado_sorted = sorted(tornado, key=lambda t: abs(t[1]), reverse=True)
t0 = r
for label, delta in tornado_sorted:
    ws.cell(row=r, column=1, value=label).font = F_BASE
    cell = ws.cell(row=r, column=2, value=round(delta))
    cell.number_format = NUM_K
    r += 1
bar = BarChart()
bar.type = "bar"
bar.title = "Sensitivität: Δ EBITDA 2031 (k€, Base)"
bar.height, bar.width = 10, 20
data = Reference(ws, min_col=2, min_row=t0, max_row=r - 1)
cats = Reference(ws, min_col=1, min_row=t0, max_row=r - 1)
bar.add_data(data)
bar.set_categories(cats)
bar.legend = None
ws.add_chart(bar, f"D{t0}")

r += 1
ws.cell(row=r, column=1, value="Lesart: Stückzahlen und COGS dominieren; Attach-Umsätze entscheiden über die Margenqualität ab Jahr 4; der Case bricht bei ASP −15 % nicht zusammen (Preissetzungs-Puffer).").font = F_NOTE

# ---------------------------------------------------------------- Returns

ws = wb.create_sheet("Returns")
ws.column_dimensions["A"].width = 52
for i in range(2, 5):
    ws.column_dimensions[get_column_letter(i)].width = 17
ws.cell(row=1, column=1, value="Investor Returns — Exit-Rechnung").font = F_TITLE
ws.cell(row=2, column=1, value="Konvention: Investor stellt den vollen Equity-Bedarf in Tranchen; Anteil = Verhandlungsannahme (Annahmen-Sheet). Terminalperspektive Jahr 7 mit Wachstums-/Margenannahmen je Szenario.").font = F_NOTE

r = 4
r = header_row(ws, r, ["Position", "Konservativ", "Base", "Upside"])
rows_def = []
r_invest = r
for ci, n in enumerate(SCEN, start=2):
    ws.cell(row=r, column=ci, value=f"={n}!B{refs[n]['r_need']}").number_format = NUM_K
ws.cell(row=r, column=1, value="Investiertes Kapital (= Kapitalbedarf brutto)").font = F_BASE
r += 1
r_rev5 = r
for ci, n in enumerate(SCEN, start=2):
    ws.cell(row=r, column=ci, value=f"={n}!{last_col}{refs[n]['r_rev']}").number_format = NUM_K
ws.cell(row=r, column=1, value="Umsatz 2031 (Jahr 5)").font = F_BASE
r += 1
r_ev5 = r
stake_col = {"Konservativ": "B", "Base": "C", "Upside": "D"}
for ci, n in enumerate(SCEN, start=2):
    col = get_column_letter(ci)
    ws.cell(row=r, column=ci, value=f"={col}{r_rev5}*Annahmen!${stake_col[n]}$17").number_format = NUM_K
ws.cell(row=r, column=1, value="Enterprise Value 2031 (Umsatz × Exit-Multiple)").font = F_BASE
r += 1
r_moic5 = r
for ci, n in enumerate(SCEN, start=2):
    col = get_column_letter(ci)
    ws.cell(row=r, column=ci, value=f"={col}{r_ev5}*Annahmen!${stake_col[n]}$18/{col}{r_invest}").number_format = NUM_X
ws.cell(row=r, column=1, value="MOIC bei Exit Jahr 5").font = F_BOLD
r += 2

# Terminalperspektive Jahr 7
ws.cell(row=r, column=1, value="Terminalperspektive Jahr 7 (2033)").font = F_H2
r += 1
r_rev7 = r
for ci, n in enumerate(SCEN, start=2):
    g = SCEN[n]["ext_growth"]
    col = get_column_letter(ci)
    ws.cell(row=r, column=ci, value=f"={col}{r_rev5}*{(1 + g) ** 2:.4f}").number_format = NUM_K
ws.cell(row=r, column=1, value=f"Umsatz 2033 (Wachstum p.a.: K {SCEN['Konservativ']['ext_growth']:.0%} / B {SCEN['Base']['ext_growth']:.0%} / U {SCEN['Upside']['ext_growth']:.0%})").font = F_BASE
r += 1
r_ev7 = r
for ci, n in enumerate(SCEN, start=2):
    col = get_column_letter(ci)
    ws.cell(row=r, column=ci, value=f"={col}{r_rev7}*Annahmen!${stake_col[n]}$17").number_format = NUM_K
ws.cell(row=r, column=1, value="Enterprise Value 2033").font = F_BASE
r += 1
r_moic7 = r
for ci, n in enumerate(SCEN, start=2):
    col = get_column_letter(ci)
    ws.cell(row=r, column=ci, value=f"={col}{r_ev7}*Annahmen!${stake_col[n]}$18/{col}{r_invest}").number_format = NUM_X
ws.cell(row=r, column=1, value="MOIC bei Exit Jahr 7").font = F_BOLD
r += 1
# Cashflow-Tabelle für IRR (Tranchen 40/35/25 % in t0-t2, Exit-Erlös in t7)
ws.cell(row=r, column=1, value="Cashflows für IRR (t0–t7)").font = F_BASE
cf_row = r + 1
tranches = [-0.40, -0.35, -0.25, 0, 0, 0, 0, None]  # None = Exit
for t in range(8):
    ws.cell(row=cf_row + t, column=1, value=f"t{t} ({2026 + t})").font = F_NOTE
    for ci, n in enumerate(SCEN, start=2):
        col = get_column_letter(ci)
        if tranches[t] is None:
            val = f"={col}{r_ev7}*Annahmen!${stake_col[n]}$18"
        else:
            val = f"={tranches[t]}*{col}{r_invest}"
        ws.cell(row=cf_row + t, column=ci, value=val).number_format = NUM_K
r = cf_row + 8
r_irr7 = r
for ci, n in enumerate(SCEN, start=2):
    col = get_column_letter(ci)
    ws.cell(row=r, column=ci, value=f"=IRR({col}{cf_row}:{col}{cf_row + 7})").number_format = NUM_PCT1
ws.cell(row=r, column=1, value="IRR (7 Jahre, tranchiert 40/35/25 %)").font = F_BOLD
r += 2
ws.cell(row=r, column=1, value="Lesart: Nur der Upside-Fall liefert venture-taugliche Renditen. Base = solide, aber unter Venture-Schwelle. Konservativ = Kapitalverlust → No-Go-Kriterien in business-plan/investment-thesis.md.").font = F_NOTE

wb.save("business-plan/financial-model.xlsx")
print("financial-model.xlsx geschrieben.")

# Kontrollausgabe der Python-Nachrechnung (muss den Excel-Formeln entsprechen)
for n in SCEN:
    e, f = simulate(n)
    print(f"{n:12s}  EBITDA 2031 ≈ {e:8.0f} k€   Kapitalbedarf ≈ {f:8.0f} k€")
